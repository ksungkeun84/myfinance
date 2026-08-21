import pandas as pd
from openpyxl import load_workbook

# 엑셀 파일 경로
file_path = "data/7145601085-01.xlsx"

def openpyxl_example():
    print("=== openpyxl 예제 ===")
    try:
        # data_only=True로 설정하면 수식 대신 계산된 최종 값을 읽어옵니다.
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active

        # 1. 엑셀 셀 주소(예: 'B2')로 직접 값 읽기
        b2_value = sheet['B2'].value
        print(f"1. sheet['B2'] 값: {b2_value}")

        # 2. 행(row)과 열(column) 번호로 값 읽기 (행은 1부터, 열은 1부터 시작)
        # B2 셀은 row=2, column=2 (A=1, B=2, C=3, ...)
        row_num = 2
        col_num = 2
        cell_value = sheet.cell(row=row_num, column=col_num).value
        print(f"2. sheet.cell(row={row_num}, column={col_num}) 값: {cell_value}")

        # 3. 모든 행을 순회하며 특정 열의 값 가져오기
        print("\n3. 행별 순회 및 값 출력 (최대 5행만 출력):")
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx > 5:
                break
            # row는 해당 행의 값들이 담긴 튜플입니다.
            # B열(인덱스 1)의 값을 출력하는 예시
            b_val = row[1] if len(row) > 1 else None
            print(f"   [행 {row_idx}] 전체 데이터: {row} | B열 값: {b_val}")

    except Exception as e:
        print(f"에러 발생: {e}")

def pandas_example():
    print("\n=== pandas 예제 ===")
    try:
        df = pd.read_excel(file_path)

        # 1. 행 인덱스(0부터 시작)와 컬럼명으로 단일 값 읽기 (.at 사용)
        # pandas는 첫 행을 헤더(컬럼명)로 사용하므로, df.at[0, '컬럼명']은 첫 번째 데이터 행의 값을 의미합니다.
        first_col_name = df.columns[1]  # 두 번째 컬럼 이름 가져오기 (예: B열 위치)
        cell_value_at = df.at[0, first_col_name]
        print(f"1. df.at[0, '{first_col_name}'] 값: {cell_value_at}")

        # 2. 행 번호와 열 번호(숫자 인덱스, 0부터 시작)로 읽기 (.iloc 사용)
        # 0번째 행, 1번째 열 (엑셀 기준 대략 B2 셀 근처 위치)
        cell_value_iloc = df.iloc[0, 1]
        print(f"2. df.iloc[0, 1] 값: {cell_value_iloc}")

        # 3. 행별로 데이터 출력하기 (최대 3행)
        print("\n3. df.iterrows()를 이용한 행별 순회 (최대 3행):")
        for index, row in df.head(3).iterrows():
            print(f"   [인덱스 {index}] {row.to_dict()}")

    except Exception as e:
        print(f"에러 발생: {e}")

if __name__ == "__main__":
    openpyxl_example()
    pandas_example()
